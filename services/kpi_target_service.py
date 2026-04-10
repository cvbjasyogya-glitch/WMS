import os
import re
from functools import lru_cache

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - handled gracefully in runtime
    load_workbook = None


DEFAULT_KPI_WORKBOOK_PATH = os.getenv(
    "KPI_TARGET_WORKBOOK_PATH",
    r"c:\Users\HYPE AMD\Downloads\TARGET & KPI.xlsx",
)

TARGET_SHEET_META = {
    "TARGET MATARAM": {
        "warehouse_key": "mataram",
        "warehouse_label": "Gudang Mataram",
    },
    "TARGET MEGA": {
        "warehouse_key": "mega",
        "warehouse_label": "Gudang Mega",
    },
    "STRINGERS": {
        "warehouse_key": "stringers",
        "warehouse_label": "Stringers",
    },
}

KPI_ASSIGNMENT_SHEET_META = {
    "KPI APRIL MATARAM": {
        "warehouse_key": "mataram",
        "warehouse_label": "Gudang Mataram",
    },
    "KPI APRIL MEGA": {
        "warehouse_key": "mega",
        "warehouse_label": "Gudang Mega",
    },
}


def _safe_text(value):
    if value is None:
        return ""
    return str(value).strip()


def _safe_float(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    safe_value = _safe_text(value).replace(",", "")
    if not safe_value:
        return None
    try:
        return float(safe_value)
    except (TypeError, ValueError):
        return None


def _normalize_lookup_name(value):
    safe_value = _safe_text(value).lower()
    safe_value = re.sub(r"[^a-z0-9]+", " ", safe_value)
    return " ".join(safe_value.split())


def _slugify(value):
    safe_value = _normalize_lookup_name(value).replace(" ", "-")
    return safe_value.strip("-")


def normalize_week_key(value):
    safe_value = _safe_text(value).upper().replace(" ", "")
    return safe_value if safe_value in {"W1", "W2", "W3", "W4"} else "W1"


def resolve_kpi_workbook_path(path=None):
    safe_path = _safe_text(path) or _safe_text(DEFAULT_KPI_WORKBOOK_PATH)
    return safe_path


def _resolve_kpi_month_label(title_text, fallback_label):
    title = _safe_text(title_text)
    match = re.search(r"\(([^)]+)\)", title)
    if match:
        return _safe_text(match.group(1)) or fallback_label
    return fallback_label


def _build_target_sheet_columns(header_groups, header_actions, header_units):
    columns = []
    last_group = ""
    for index in range(1, max(len(header_groups), len(header_actions), len(header_units))):
        group_text = _safe_text(header_groups[index] if index < len(header_groups) else "")
        action_text = _safe_text(header_actions[index] if index < len(header_actions) else "")
        unit_text = _safe_text(header_units[index] if index < len(header_units) else "")

        if group_text:
            last_group = group_text
        elif last_group and (action_text or unit_text):
            group_text = last_group

        label_parts = [part for part in [group_text, action_text, unit_text] if part]
        if not label_parts:
            continue

        label = " / ".join(label_parts)
        columns.append(
            {
                "index": index,
                "group_label": group_text or "-",
                "action_label": action_text or "-",
                "unit_label": unit_text or "-",
                "label": label,
                "item_key": _slugify(label) or f"col-{index}",
            }
        )
    return columns


def _parse_target_sheet(sheet, meta):
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 5:
        return {
            "sheet_name": sheet.title,
            "warehouse_key": meta["warehouse_key"],
            "warehouse_label": meta["warehouse_label"],
            "title": sheet.title,
            "columns": [],
            "staff_rows": [],
        }

    title = _safe_text(rows[0][0] if rows[0] else "") or sheet.title
    columns = _build_target_sheet_columns(rows[1], rows[2], rows[3])
    staff_rows = []

    for raw_row in rows[4:]:
        if not raw_row or not any(_safe_text(value) for value in raw_row):
            continue
        staff_name = _safe_text(raw_row[0] if len(raw_row) >= 1 else "")
        if not staff_name:
            continue

        metrics = []
        for column in columns:
            value = raw_row[column["index"]] if column["index"] < len(raw_row) else None
            metrics.append(
                {
                    "item_key": column["item_key"],
                    "group_label": column["group_label"],
                    "action_label": column["action_label"],
                    "unit_label": column["unit_label"],
                    "label": column["label"],
                    "value": _safe_float(value),
                    "display_value": _safe_text(value) or "-",
                }
            )

        staff_rows.append(
            {
                "staff_name": staff_name,
                "lookup_name": _normalize_lookup_name(staff_name),
                "warehouse_key": meta["warehouse_key"],
                "warehouse_label": meta["warehouse_label"],
                "metrics": metrics,
            }
        )

    return {
        "sheet_name": sheet.title,
        "warehouse_key": meta["warehouse_key"],
        "warehouse_label": meta["warehouse_label"],
        "title": title,
        "columns": columns,
        "staff_rows": staff_rows,
    }


def _parse_kpi_assignment_sheet(sheet, meta):
    rows = list(sheet.iter_rows(values_only=True))
    assignments = []
    index = 0
    fallback_month = sheet.title.replace("KPI", "").strip() or "Bulan Aktif"

    while index < len(rows):
        first_cell = _safe_text(rows[index][0] if rows[index] else "")
        if not first_cell.lower().startswith("target individu kpi"):
            index += 1
            continue

        month_label = _resolve_kpi_month_label(first_cell, fallback_month)
        employee_name = _safe_text(rows[index + 1][0] if index + 1 < len(rows) and rows[index + 1] else "")
        if not employee_name:
            index += 1
            continue

        item_index = index + 4
        current_group = ""
        pass_threshold = None
        items = []

        while item_index < len(rows):
            current_row = rows[item_index]
            first_value = _safe_text(current_row[0] if len(current_row) >= 1 else "")
            second_value = _safe_text(current_row[1] if len(current_row) >= 2 else "")

            if first_value.lower().startswith("target individu kpi"):
                break

            if first_value.lower().startswith("belum lulus"):
                pass_threshold = _safe_float(current_row[6] if len(current_row) >= 7 else None)
                item_index += 1
                continue

            if not any(_safe_text(value) for value in current_row):
                next_value = ""
                if item_index + 1 < len(rows) and rows[item_index + 1]:
                    next_value = _safe_text(rows[item_index + 1][0] if len(rows[item_index + 1]) >= 1 else "")
                if next_value.lower().startswith("target individu kpi"):
                    break
                item_index += 1
                continue

            if second_value:
                if first_value:
                    current_group = first_value
                target_value = _safe_float(current_row[2] if len(current_row) >= 3 else None)
                weight_value = _safe_float(current_row[9] if len(current_row) >= 10 else None)
                total_target_value = _safe_float(current_row[10] if len(current_row) >= 11 else None)
                items.append(
                    {
                        "item_key": _slugify(
                            f"{sheet.title}-{employee_name}-{current_group or 'metric'}-{second_value}"
                        )
                        or f"item-{len(items) + 1}",
                        "group_label": current_group or first_value or "KPI",
                        "metric_name": second_value,
                        "unit_label": _safe_text(current_row[1] if len(current_row) >= 2 else ""),
                        "target_value": target_value,
                        "weight_value": weight_value,
                        "total_target_value": total_target_value,
                        "week_defaults": {
                            "W1": _safe_float(current_row[3] if len(current_row) >= 4 else None),
                            "W2": _safe_float(current_row[4] if len(current_row) >= 5 else None),
                            "W3": _safe_float(current_row[5] if len(current_row) >= 6 else None),
                            "W4": _safe_float(current_row[6] if len(current_row) >= 7 else None),
                        },
                    }
                )
            item_index += 1

        assignments.append(
            {
                "assignment_key": _slugify(f"{sheet.title}-{employee_name}") or f"assignment-{len(assignments) + 1}",
                "assignment_name": f"{employee_name} - {month_label}",
                "employee_name": employee_name,
                "lookup_name": _normalize_lookup_name(employee_name),
                "warehouse_key": meta["warehouse_key"],
                "warehouse_label": meta["warehouse_label"],
                "month_label": month_label,
                "pass_threshold": pass_threshold,
                "items": items,
            }
        )
        index = item_index

    return assignments


def _parse_define_sheet(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    items = []
    current_group = ""
    for raw_row in rows[2:]:
        first_value = _safe_text(raw_row[0] if len(raw_row) >= 1 else "")
        second_value = _safe_text(raw_row[1] if len(raw_row) >= 2 else "")
        if not first_value and not second_value:
            continue
        if first_value.lower().startswith("nilai realisasi") or first_value.lower().startswith("excellent"):
            continue
        if first_value:
            current_group = first_value
        items.append(
            {
                "group_label": current_group or first_value or "Target",
                "metric_name": second_value or "-",
                "description": _safe_text(raw_row[2] if len(raw_row) >= 3 else ""),
                "weight_value": _safe_float(raw_row[3] if len(raw_row) >= 4 else None),
                "total_weight_value": _safe_float(raw_row[5] if len(raw_row) >= 6 else None),
            }
        )
    return items


@lru_cache(maxsize=4)
def _load_snapshot_cached(workbook_path, workbook_mtime):
    snapshot = {
        "path": workbook_path,
        "exists": False,
        "title": "KPI & Target",
        "target_sections": [],
        "assignments": [],
        "definitions": [],
        "minimum_rules": [],
        "errors": [],
    }

    if not workbook_path or not os.path.exists(workbook_path):
        snapshot["errors"].append("Workbook KPI belum ditemukan.")
        return snapshot

    if load_workbook is None:
        snapshot["errors"].append("Dependensi openpyxl belum tersedia.")
        return snapshot

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    snapshot["exists"] = True

    for sheet_name, meta in TARGET_SHEET_META.items():
        if sheet_name in workbook.sheetnames:
            snapshot["target_sections"].append(_parse_target_sheet(workbook[sheet_name], meta))

    for sheet_name, meta in KPI_ASSIGNMENT_SHEET_META.items():
        if sheet_name in workbook.sheetnames:
            snapshot["assignments"].extend(_parse_kpi_assignment_sheet(workbook[sheet_name], meta))

    if "DEFINE" in workbook.sheetnames:
        snapshot["definitions"] = _parse_define_sheet(workbook["DEFINE"])

    if "DEFINE 2" in workbook.sheetnames:
        minimum_rules = []
        rows = list(workbook["DEFINE 2"].iter_rows(values_only=True))
        current_group = ""
        for raw_row in rows[2:]:
            first_value = _safe_text(raw_row[0] if len(raw_row) >= 1 else "")
            second_value = _safe_text(raw_row[1] if len(raw_row) >= 2 else "")
            if not first_value and not second_value:
                continue
            if first_value.lower().startswith("nilai realisasi") or first_value.lower().startswith("excellent"):
                continue
            if first_value:
                current_group = first_value
            minimum_rules.append(
                {
                    "group_label": current_group or first_value or "Target",
                    "metric_name": second_value or "-",
                    "minimum_score": _safe_float(raw_row[3] if len(raw_row) >= 4 else None),
                }
            )
        snapshot["minimum_rules"] = minimum_rules

    return snapshot


def load_kpi_target_snapshot(path=None):
    workbook_path = resolve_kpi_workbook_path(path)
    workbook_mtime = 0
    if workbook_path and os.path.exists(workbook_path):
        workbook_mtime = int(os.path.getmtime(workbook_path))
    return _load_snapshot_cached(workbook_path, workbook_mtime)


def _resolve_warehouse_key_from_label(warehouse_label):
    safe_label = _normalize_lookup_name(warehouse_label)
    if "mega" in safe_label:
        return "mega"
    if "stringer" in safe_label:
        return "stringers"
    if "mataram" in safe_label:
        return "mataram"
    return ""


def get_assignment_options(snapshot, warehouse_label=None):
    warehouse_key = _resolve_warehouse_key_from_label(warehouse_label)
    assignments = snapshot.get("assignments", [])
    if warehouse_key:
        filtered = [item for item in assignments if item.get("warehouse_key") == warehouse_key]
        if filtered:
            assignments = filtered
    return assignments


def find_staff_assignment(snapshot, employee_name=None, warehouse_label=None, username=None):
    assignments = get_assignment_options(snapshot, warehouse_label)
    lookup_candidates = [
        _normalize_lookup_name(employee_name),
        _normalize_lookup_name(username),
    ]
    lookup_candidates = [value for value in lookup_candidates if value]
    if not lookup_candidates:
        return None

    for candidate in lookup_candidates:
        exact_match = next(
            (item for item in assignments if item.get("lookup_name") == candidate),
            None,
        )
        if exact_match:
            return exact_match

    for candidate in lookup_candidates:
        fuzzy_match = next(
            (
                item
                for item in assignments
                if candidate in item.get("lookup_name", "")
                or item.get("lookup_name", "") in candidate
            ),
            None,
        )
        if fuzzy_match:
            return fuzzy_match

    return None
