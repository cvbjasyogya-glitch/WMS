import re
import warnings
from collections import defaultdict
from datetime import date, datetime
from io import BytesIO

from openpyxl import load_workbook


CRM_IMPORT_SOURCE_PREFIX = "crm_excel:"
CRM_IMPORT_NOTE_PREFIX = "Import CRM Excel"
DATABASE_SHEET_NAME = "DATABASE"
SALES_SHEET_NAME = "DATABASE PENJUALAN"
MEMBER_SHEETS = (
    ("MEMBER SENARAN MEGA", "mega"),
    ("MEMBER SENARAN MTRM", "mataram"),
    ("MEMBER SENARAN MEGA TENIS", "mega"),
)
STRINGING_REWARD_DEFAULT = 75000


def _safe_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_space(value):
    return re.sub(r"\s+", " ", _safe_text(value)).strip()


def _normalize_key(value):
    return re.sub(r"[^a-z0-9]+", "", _safe_text(value).lower())


def _normalize_name(value):
    return _normalize_space(value).upper()


def _normalize_phone(value):
    text = _safe_text(value)
    if not text:
        return ""
    lowered = text.lower()
    for prefix in ("https://", "http://"):
        if lowered.startswith(prefix):
            text = text[len(prefix):]
            lowered = text.lower()
    if lowered.startswith("wa.me/"):
        text = text[6:]

    digits = re.sub(r"\D+", "", text)
    if not digits:
        return ""
    if digits.startswith("620"):
        digits = "62" + digits[3:]
    elif digits.startswith("0"):
        digits = "62" + digits[1:]
    elif digits.startswith("8"):
        digits = "62" + digits
    if len(digits) < 9:
        return ""
    return digits


def _normalize_invoice(value):
    text = _safe_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+(\.0+)?", text):
        return str(int(float(text)))
    return _normalize_space(text).upper()


def _normalize_excel_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    text = _safe_text(value)
    if not text or text.startswith("#"):
        return None

    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _to_int(value, default=0):
    try:
        if value in (None, ""):
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _merge_notes(existing, incoming):
    existing_text = _normalize_space(existing)
    incoming_text = _normalize_space(incoming)
    if not existing_text:
        return incoming_text or None
    if not incoming_text:
        return existing_text
    if incoming_text in existing_text:
        return existing_text
    if existing_text in incoming_text:
        return incoming_text
    return f"{existing_text}\n{incoming_text}"


def _earliest_iso_date(*values):
    valid_values = sorted(value for value in values if _normalize_space(value))
    return valid_values[0] if valid_values else None


def _build_import_source(filename):
    safe_name = _normalize_space(filename) or "crm-import.xlsx"
    return f"{CRM_IMPORT_SOURCE_PREFIX}{safe_name}"


def _map_customer_type(raw_category, *, force_member=False):
    if force_member:
        return "member"

    category_key = _normalize_key(raw_category)
    if any(token in category_key for token in ("instansi", "pemerintahan", "sekolah", "cv")):
        return "wholesale"
    return "retail"


def _format_customer_note(row):
    note_parts = []
    if _normalize_space(row.get("category")):
        note_parts.append(f"Kategori file: {_normalize_space(row.get('category'))}")
    if _normalize_space(row.get("gender")):
        note_parts.append(f"Gender file: {_normalize_space(row.get('gender'))}")
    if _normalize_space(row.get("status")):
        note_parts.append(f"Status file: {_normalize_space(row.get('status'))}")
    warehouse_label = _normalize_space(row.get("warehouse_label"))
    if warehouse_label:
        note_parts.append(f"Lokasi file: {warehouse_label}")
    if not note_parts:
        return None
    return f"{CRM_IMPORT_NOTE_PREFIX} customer. " + " | ".join(note_parts)


def _build_purchase_note(row, items_summary):
    note_parts = [f"{CRM_IMPORT_NOTE_PREFIX} histori pembelian."]
    category = _normalize_space(row.get("category"))
    if category:
        note_parts.append(f"Kategori: {category}")
    location = _normalize_space(row.get("warehouse_label"))
    if location:
        note_parts.append(f"Lokasi file: {location}")
    if items_summary:
        note_parts.append(f"Ringkasan item: {items_summary}")
    return " | ".join(note_parts)


def _build_member_note(member_row):
    note_parts = [
        f"{CRM_IMPORT_NOTE_PREFIX} member dari sheet {member_row['source_sheet']}.",
        f"ID file: {member_row['raw_member_id'] or '-'}",
    ]
    if member_row["visit_refs"]:
        note_parts.append("Riwayat kunjungan: " + ", ".join(member_row["visit_refs"]))
    if member_row["free_marker"]:
        note_parts.append(f"Kolom FREE: {member_row['free_marker']}")
    if member_row["extra_notes"]:
        note_parts.append("Catatan file: " + ", ".join(member_row["extra_notes"]))
    return " | ".join(note_parts)


def _build_member_code(warehouse_hint, raw_member_id, source_sheet):
    sanitized_member_id = re.sub(r"[^A-Z0-9]+", "-", _safe_text(raw_member_id).upper()).strip("-") or "UNKNOWN"
    warehouse_segment = "MTRM" if warehouse_hint == "mataram" else "MEGA"
    if "TENIS" in _safe_text(source_sheet).upper():
        warehouse_segment = f"{warehouse_segment}-TENIS"
    return f"CRM-STR-{warehouse_segment}-{sanitized_member_id}"


class WarehouseResolver:
    def __init__(self, warehouse_rows):
        self._rows = [dict(row) for row in warehouse_rows]
        self._by_exact_key = {}
        self._by_keyword = {}
        for row in self._rows:
            warehouse_id = int(row["id"])
            normalized_name = _normalize_key(row.get("name"))
            if normalized_name:
                self._by_exact_key[normalized_name] = warehouse_id
            if "mataram" in normalized_name:
                self._by_keyword["mataram"] = warehouse_id
                self._by_keyword["mtrm"] = warehouse_id
            if "mega" in normalized_name:
                self._by_keyword["mega"] = warehouse_id

    def resolve(self, raw_label):
        normalized = _normalize_key(raw_label)
        if not normalized:
            return None
        if normalized in self._by_exact_key:
            return self._by_exact_key[normalized]
        if "mataram" in normalized or "mtrm" in normalized:
            return self._by_keyword.get("mataram") or self._by_keyword.get("mtrm")
        if "mega" in normalized:
            return self._by_keyword.get("mega")
        for known_key, warehouse_id in self._by_exact_key.items():
            if normalized in known_key or known_key in normalized:
                return warehouse_id
        return None


def _parse_customer_rows(workbook):
    if DATABASE_SHEET_NAME not in workbook.sheetnames:
        raise ValueError("Sheet DATABASE tidak ditemukan di file CRM.")

    rows = []
    worksheet = workbook[DATABASE_SHEET_NAME]
    for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        customer_name = _normalize_space(row[1] if len(row) > 1 else None)
        invoice_no = _normalize_invoice(row[4] if len(row) > 4 else None)
        warehouse_label = _normalize_space(row[6] if len(row) > 6 else None)
        if not any([customer_name, invoice_no, warehouse_label]):
            continue

        rows.append(
            {
                "row_index": row_index,
                "purchase_date": _normalize_excel_date(row[0] if len(row) > 0 else None),
                "customer_name": customer_name,
                "gender": _normalize_space(row[2] if len(row) > 2 else None),
                "category": _normalize_space(row[3] if len(row) > 3 else None),
                "invoice_no": invoice_no,
                "phone": _normalize_phone(row[5] if len(row) > 5 else None),
                "raw_phone": _normalize_space(row[5] if len(row) > 5 else None),
                "warehouse_label": warehouse_label,
                "status": _normalize_space(row[7] if len(row) > 7 else None)
                or _normalize_space(row[8] if len(row) > 8 else None),
            }
        )
    return rows


def _parse_sales_rows(workbook):
    if SALES_SHEET_NAME not in workbook.sheetnames:
        return {}, {}

    sales_by_invoice = defaultdict(list)
    sales_by_invoice_and_location = defaultdict(list)

    worksheet = workbook[SALES_SHEET_NAME]
    current_invoice = ""
    current_location = ""
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        invoice_no = _normalize_invoice(row[0] if len(row) > 0 else None) or current_invoice
        item_name = _normalize_space(row[1] if len(row) > 1 else None)
        location = _normalize_space(row[2] if len(row) > 2 else None) or current_location
        qty = _to_int(row[3] if len(row) > 3 else None, 0)

        if invoice_no:
            current_invoice = invoice_no
        if location:
            current_location = location

        if not invoice_no or not item_name:
            continue

        item_row = {
            "item_name": item_name,
            "qty": max(1, qty or 1),
        }
        sales_by_invoice[invoice_no].append(item_row)
        sales_by_invoice_and_location[(_normalize_key(location), invoice_no)].append(item_row)

    return dict(sales_by_invoice), dict(sales_by_invoice_and_location)


def _parse_member_rows(workbook):
    member_rows = []
    for sheet_name, warehouse_hint in MEMBER_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue

        worksheet = workbook[sheet_name]
        for row_index, row in enumerate(worksheet.iter_rows(min_row=4, values_only=True), start=4):
            raw_member_id = _normalize_space(row[0] if len(row) > 0 else None)
            customer_name = _normalize_space(row[1] if len(row) > 1 else None)
            if not raw_member_id and not customer_name:
                continue

            visit_refs = []
            for visit_cell in row[3:9]:
                normalized_visit = _normalize_invoice(visit_cell)
                if normalized_visit:
                    visit_refs.append(normalized_visit)
            free_marker = _normalize_space(row[9] if len(row) > 9 else None)
            extra_notes = [_normalize_space(cell) for cell in row[10:] if _normalize_space(cell)]

            member_rows.append(
                {
                    "row_index": row_index,
                    "source_sheet": sheet_name,
                    "warehouse_hint": warehouse_hint,
                    "raw_member_id": raw_member_id,
                    "customer_name": customer_name,
                    "join_date": _normalize_excel_date(row[2] if len(row) > 2 else None),
                    "visit_refs": visit_refs,
                    "opening_stringing_visits": len(visit_refs),
                    "opening_reward_redeemed": 1 if free_marker else 0,
                    "free_marker": free_marker,
                    "extra_notes": extra_notes,
                }
            )
    return member_rows


def _load_customer_cache(db, allowed_warehouse_ids=None):
    params = []
    query = """
        SELECT
            id,
            warehouse_id,
            customer_name,
            phone,
            customer_type,
            marketing_channel,
            note,
            contact_person
        FROM crm_customers
        WHERE 1=1
    """
    if allowed_warehouse_ids:
        placeholders = ",".join("?" for _ in allowed_warehouse_ids)
        query += f" AND warehouse_id IN ({placeholders})"
        params.extend(allowed_warehouse_ids)

    rows = [dict(row) for row in db.execute(query, params).fetchall()]
    by_phone = {}
    by_name = {}
    for row in rows:
        warehouse_id = int(row["warehouse_id"])
        normalized_phone = _normalize_phone(row.get("phone"))
        normalized_name = _normalize_name(row.get("customer_name"))
        if normalized_phone:
            by_phone[(warehouse_id, normalized_phone)] = row
        if normalized_name and (warehouse_id, normalized_name) not in by_name:
            by_name[(warehouse_id, normalized_name)] = row
    return by_phone, by_name


def _load_purchase_cache(db, allowed_warehouse_ids=None):
    params = []
    query = """
        SELECT
            id,
            customer_id,
            warehouse_id,
            purchase_date,
            invoice_no,
            channel,
            transaction_type,
            items_count,
            total_amount,
            note,
            handled_by,
            import_source,
            import_items_summary,
            import_total_qty
        FROM crm_purchase_records
        WHERE TRIM(COALESCE(invoice_no, '')) <> ''
    """
    if allowed_warehouse_ids:
        placeholders = ",".join("?" for _ in allowed_warehouse_ids)
        query += f" AND warehouse_id IN ({placeholders})"
        params.extend(allowed_warehouse_ids)

    rows = [dict(row) for row in db.execute(query, params).fetchall()]
    return {
        (int(row["warehouse_id"]), _normalize_invoice(row.get("invoice_no"))): row
        for row in rows
        if _normalize_invoice(row.get("invoice_no"))
    }


def _load_membership_cache(db, allowed_warehouse_ids=None):
    params = []
    query = """
        SELECT
            id,
            customer_id,
            warehouse_id,
            member_code,
            member_type,
            tier,
            status,
            join_date,
            expiry_date,
            points,
            requested_by_staff_id,
            reward_unit_amount,
            opening_stringing_visits,
            opening_reward_redeemed,
            benefit_note,
            note
        FROM crm_memberships
        WHERE 1=1
    """
    if allowed_warehouse_ids:
        placeholders = ",".join("?" for _ in allowed_warehouse_ids)
        query += f" AND warehouse_id IN ({placeholders})"
        params.extend(allowed_warehouse_ids)

    rows = [dict(row) for row in db.execute(query, params).fetchall()]
    by_customer = {}
    by_member_code = {}
    for row in rows:
        by_customer[int(row["customer_id"])] = row
        by_member_code[_safe_text(row.get("member_code")).upper()] = row
    return by_customer, by_member_code


def _upsert_customer(db, cache_by_phone, cache_by_name, warehouse_id, customer_name, *, phone="", customer_type="retail", note=None):
    normalized_name = _normalize_name(customer_name)
    normalized_phone = _normalize_phone(phone)
    existing = None
    if normalized_phone:
        existing = cache_by_phone.get((warehouse_id, normalized_phone))
    if not existing and normalized_name:
        existing = cache_by_name.get((warehouse_id, normalized_name))

    if existing:
        updates = {}
        if normalized_phone and not _normalize_phone(existing.get("phone")):
            updates["phone"] = normalized_phone
        if customer_type == "member" and (existing.get("customer_type") or "retail") != "member":
            updates["customer_type"] = "member"
        merged_note = _merge_notes(existing.get("note"), note)
        if merged_note != _normalize_space(existing.get("note")):
            updates["note"] = merged_note
        if not _normalize_space(existing.get("marketing_channel")):
            updates["marketing_channel"] = "Import CRM Excel"
        if updates:
            assignments = ", ".join(f"{column}=?" for column in updates)
            db.execute(
                f"UPDATE crm_customers SET {assignments}, updated_at=CURRENT_TIMESTAMP WHERE id=?",
                [*updates.values(), existing["id"]],
            )
            existing = {**existing, **updates}
            if normalized_phone:
                cache_by_phone[(warehouse_id, normalized_phone)] = existing
            cache_by_name[(warehouse_id, normalized_name)] = existing
            return existing, False, True
        return existing, False, False

    cursor = db.execute(
        """
        INSERT INTO crm_customers(
            warehouse_id,
            customer_name,
            phone,
            customer_type,
            marketing_channel,
            note
        )
        VALUES (?,?,?,?,?,?)
        """,
        (
            warehouse_id,
            _normalize_space(customer_name),
            normalized_phone or None,
            customer_type,
            "Import CRM Excel",
            note or None,
        ),
    )
    created = {
        "id": cursor.lastrowid,
        "warehouse_id": warehouse_id,
        "customer_name": _normalize_space(customer_name),
        "phone": normalized_phone or None,
        "customer_type": customer_type,
        "marketing_channel": "Import CRM Excel",
        "note": note or None,
        "contact_person": None,
    }
    if normalized_phone:
        cache_by_phone[(warehouse_id, normalized_phone)] = created
    cache_by_name[(warehouse_id, normalized_name)] = created
    return created, True, False


def _upsert_purchase(
    db,
    purchase_cache,
    *,
    warehouse_id,
    customer_id,
    purchase_date,
    invoice_no,
    handled_by,
    import_source,
    import_items_summary,
    import_total_qty,
    note,
):
    cache_key = (warehouse_id, invoice_no)
    existing = purchase_cache.get(cache_key)
    if existing and not _safe_text(existing.get("import_source")).startswith(CRM_IMPORT_SOURCE_PREFIX):
        return "conflict"

    values = {
        "customer_id": customer_id,
        "member_id": None,
        "warehouse_id": warehouse_id,
        "purchase_date": purchase_date,
        "invoice_no": invoice_no,
        "channel": "store",
        "transaction_type": "purchase",
        "items_count": import_total_qty,
        "total_amount": 0,
        "note": note or None,
        "handled_by": handled_by,
        "import_source": import_source,
        "import_items_summary": import_items_summary or None,
        "import_total_qty": import_total_qty,
    }

    if existing:
        db.execute(
            """
            UPDATE crm_purchase_records
            SET
                customer_id=?,
                member_id=?,
                warehouse_id=?,
                purchase_date=?,
                invoice_no=?,
                channel=?,
                transaction_type=?,
                items_count=?,
                total_amount=?,
                note=?,
                handled_by=?,
                import_source=?,
                import_items_summary=?,
                import_total_qty=?,
                updated_at=CURRENT_TIMESTAMP
            WHERE id=?
            """,
            (
                values["customer_id"],
                values["member_id"],
                values["warehouse_id"],
                values["purchase_date"],
                values["invoice_no"],
                values["channel"],
                values["transaction_type"],
                values["items_count"],
                values["total_amount"],
                values["note"],
                values["handled_by"],
                values["import_source"],
                values["import_items_summary"],
                values["import_total_qty"],
                existing["id"],
            ),
        )
        purchase_cache[cache_key] = {**existing, **values}
        return "updated"

    cursor = db.execute(
        """
        INSERT INTO crm_purchase_records(
            customer_id,
            member_id,
            warehouse_id,
            purchase_date,
            invoice_no,
            channel,
            transaction_type,
            items_count,
            total_amount,
            note,
            handled_by,
            import_source,
            import_items_summary,
            import_total_qty
        )
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            values["customer_id"],
            values["member_id"],
            values["warehouse_id"],
            values["purchase_date"],
            values["invoice_no"],
            values["channel"],
            values["transaction_type"],
            values["items_count"],
            values["total_amount"],
            values["note"],
            values["handled_by"],
            values["import_source"],
            values["import_items_summary"],
            values["import_total_qty"],
        ),
    )
    purchase_cache[cache_key] = {**values, "id": cursor.lastrowid}
    return "created"


def _upsert_membership(
    db,
    by_customer,
    by_member_code,
    *,
    customer_id,
    warehouse_id,
    member_code,
    join_date,
    opening_stringing_visits,
    opening_reward_redeemed,
    benefit_note,
    note,
):
    existing = by_customer.get(customer_id) or by_member_code.get(member_code)
    if existing:
        existing_member_type = _safe_text(existing.get("member_type")).lower() or "purchase"
        existing_note = _safe_text(existing.get("note"))
        if existing_member_type not in {"stringing", ""} and CRM_IMPORT_NOTE_PREFIX not in existing_note:
            return "conflict"

        merged_note = _merge_notes(existing.get("note"), note)
        merged_benefit_note = _merge_notes(existing.get("benefit_note"), benefit_note)
        updated_join_date = _earliest_iso_date(existing.get("join_date"), join_date) or join_date
        updated_opening_visits = max(_to_int(existing.get("opening_stringing_visits"), 0), opening_stringing_visits)
        updated_reward_redeemed = max(
            _to_int(existing.get("opening_reward_redeemed"), 0),
            opening_reward_redeemed,
        )
        db.execute(
            """
            UPDATE crm_memberships
            SET
                customer_id=?,
                warehouse_id=?,
                member_code=?,
                member_type='stringing',
                tier='regular',
                status='active',
                join_date=?,
                reward_unit_amount=?,
                opening_stringing_visits=?,
                opening_reward_redeemed=?,
                benefit_note=?,
                note=?,
                updated_at=CURRENT_TIMESTAMP
            WHERE id=?
            """,
            (
                customer_id,
                warehouse_id,
                member_code,
                updated_join_date,
                STRINGING_REWARD_DEFAULT,
                updated_opening_visits,
                updated_reward_redeemed,
                merged_benefit_note,
                merged_note,
                existing["id"],
            ),
        )
        updated_row = {
            **existing,
            "customer_id": customer_id,
            "warehouse_id": warehouse_id,
            "member_code": member_code,
            "member_type": "stringing",
            "tier": "regular",
            "status": "active",
            "join_date": updated_join_date,
            "reward_unit_amount": STRINGING_REWARD_DEFAULT,
            "opening_stringing_visits": updated_opening_visits,
            "opening_reward_redeemed": updated_reward_redeemed,
            "benefit_note": merged_benefit_note,
            "note": merged_note,
        }
        by_customer[customer_id] = updated_row
        by_member_code[member_code] = updated_row
        return "updated"

    cursor = db.execute(
        """
        INSERT INTO crm_memberships(
            customer_id,
            warehouse_id,
            member_code,
            member_type,
            tier,
            status,
            join_date,
            reward_unit_amount,
            opening_stringing_visits,
            opening_reward_redeemed,
            benefit_note,
            note
        )
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """,
        (
            customer_id,
            warehouse_id,
            member_code,
            "stringing",
            "regular",
            "active",
            join_date,
            STRINGING_REWARD_DEFAULT,
            opening_stringing_visits,
            opening_reward_redeemed,
            benefit_note or None,
            note or None,
        ),
    )
    created_row = {
        "id": cursor.lastrowid,
        "customer_id": customer_id,
        "warehouse_id": warehouse_id,
        "member_code": member_code,
        "member_type": "stringing",
        "tier": "regular",
        "status": "active",
        "join_date": join_date,
        "reward_unit_amount": STRINGING_REWARD_DEFAULT,
        "opening_stringing_visits": opening_stringing_visits,
        "opening_reward_redeemed": opening_reward_redeemed,
        "benefit_note": benefit_note or None,
        "note": note or None,
    }
    by_customer[customer_id] = created_row
    by_member_code[member_code] = created_row
    return "created"


def _build_item_summary(item_rows):
    if not item_rows:
        return "", 0
    total_qty = 0
    parts = []
    for item_row in item_rows:
        qty = max(1, _to_int(item_row.get("qty"), 1))
        total_qty += qty
        parts.append(f"{_normalize_space(item_row.get('item_name'))} x{qty}")
    return " | ".join(parts), total_qty


def import_crm_workbook(db, workbook_bytes, *, filename, selected_warehouse_id=None, handled_by=None):
    if not workbook_bytes:
        raise ValueError("File CRM kosong atau gagal dibaca.")

    import_source = _build_import_source(filename)

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)

    try:
        warehouse_rows = db.execute("SELECT id, name FROM warehouses ORDER BY id ASC").fetchall()
        if not warehouse_rows:
            raise ValueError("Master gudang belum tersedia. CRM import belum bisa dijalankan.")

        resolver = WarehouseResolver(warehouse_rows)
        db.execute("PRAGMA busy_timeout = 300000")

        customer_rows = _parse_customer_rows(workbook)
        sales_by_invoice, sales_by_invoice_and_location = _parse_sales_rows(workbook)
        member_rows = _parse_member_rows(workbook)

        allowed_warehouse_ids = None
        if selected_warehouse_id:
            allowed_warehouse_ids = [int(selected_warehouse_id)]

        customer_cache_by_phone, customer_cache_by_name = _load_customer_cache(db, allowed_warehouse_ids)
        purchase_cache = _load_purchase_cache(db, allowed_warehouse_ids)
        membership_by_customer, membership_by_member_code = _load_membership_cache(db, allowed_warehouse_ids)

        summary = {
            "customer_rows": len(customer_rows),
            "member_rows": len(member_rows),
            "customers_created": 0,
            "customers_updated": 0,
            "purchases_created": 0,
            "purchases_updated": 0,
            "purchases_skipped_conflict": 0,
            "memberships_created": 0,
            "memberships_updated": 0,
            "memberships_skipped_conflict": 0,
            "skipped_unknown_warehouse": 0,
            "skipped_missing_invoice": 0,
            "processed_warehouses": set(),
        }

        customer_ids_by_key = {}
        default_join_dates = {}

        for row in customer_rows:
            warehouse_id = resolver.resolve(row.get("warehouse_label"))
            if not warehouse_id:
                summary["skipped_unknown_warehouse"] += 1
                continue
            if selected_warehouse_id and int(selected_warehouse_id) != warehouse_id:
                continue
            if not _normalize_space(row.get("customer_name")):
                continue

            customer, created, updated = _upsert_customer(
                db,
                customer_cache_by_phone,
                customer_cache_by_name,
                warehouse_id,
                row["customer_name"],
                phone=row.get("phone"),
                customer_type=_map_customer_type(row.get("category")),
                note=_format_customer_note(row),
            )
            if created:
                summary["customers_created"] += 1
            elif updated:
                summary["customers_updated"] += 1

            summary["processed_warehouses"].add(warehouse_id)
            customer_ids_by_key[(warehouse_id, row.get("invoice_no"))] = customer["id"]
            customer_ids_by_key[(warehouse_id, _normalize_name(row.get("customer_name")))] = customer["id"]

            date_key = (warehouse_id, _normalize_name(row.get("customer_name")))
            existing_default_date = default_join_dates.get(date_key)
            default_join_dates[date_key] = _earliest_iso_date(existing_default_date, row.get("purchase_date")) or row.get("purchase_date")

        db.commit()

        for member_row in member_rows:
            warehouse_id = resolver.resolve(member_row.get("warehouse_hint"))
            if not warehouse_id:
                summary["skipped_unknown_warehouse"] += 1
                continue
            if selected_warehouse_id and int(selected_warehouse_id) != warehouse_id:
                continue

            customer_name = _normalize_space(member_row.get("customer_name"))
            if not customer_name:
                continue

            customer, created, updated = _upsert_customer(
                db,
                customer_cache_by_phone,
                customer_cache_by_name,
                warehouse_id,
                customer_name,
                phone="",
                customer_type="member",
                note=f"{CRM_IMPORT_NOTE_PREFIX} member senaran dari sheet {member_row['source_sheet']}.",
            )
            if created:
                summary["customers_created"] += 1
            elif updated:
                summary["customers_updated"] += 1

            summary["processed_warehouses"].add(warehouse_id)
            member_code = _build_member_code(
                member_row["warehouse_hint"],
                member_row.get("raw_member_id"),
                member_row.get("source_sheet"),
            )
            join_date = member_row.get("join_date") or default_join_dates.get((warehouse_id, _normalize_name(customer_name)))
            if not join_date:
                join_date = date.today().isoformat()

            member_result = _upsert_membership(
                db,
                membership_by_customer,
                membership_by_member_code,
                customer_id=customer["id"],
                warehouse_id=warehouse_id,
                member_code=member_code,
                join_date=join_date,
                opening_stringing_visits=max(0, member_row.get("opening_stringing_visits") or 0),
                opening_reward_redeemed=max(0, member_row.get("opening_reward_redeemed") or 0),
                benefit_note=f"Source sheet: {member_row['source_sheet']}",
                note=_build_member_note(member_row),
            )
            if member_result == "created":
                summary["memberships_created"] += 1
            elif member_result == "updated":
                summary["memberships_updated"] += 1
            else:
                summary["memberships_skipped_conflict"] += 1

        db.commit()

        for row in customer_rows:
            warehouse_id = resolver.resolve(row.get("warehouse_label"))
            if not warehouse_id:
                continue
            if selected_warehouse_id and int(selected_warehouse_id) != warehouse_id:
                continue

            invoice_no = row.get("invoice_no")
            if not invoice_no:
                summary["skipped_missing_invoice"] += 1
                continue

            customer_id = customer_ids_by_key.get((warehouse_id, invoice_no))
            if not customer_id:
                customer_id = customer_ids_by_key.get((warehouse_id, _normalize_name(row.get("customer_name"))))
            if not customer_id:
                summary["skipped_missing_invoice"] += 1
                continue

            sales_rows = sales_by_invoice_and_location.get((_normalize_key(row.get("warehouse_label")), invoice_no))
            if not sales_rows:
                sales_rows = sales_by_invoice.get(invoice_no, [])
            items_summary, import_total_qty = _build_item_summary(sales_rows)
            purchase_date = row.get("purchase_date") or date.today().isoformat()

            purchase_result = _upsert_purchase(
                db,
                purchase_cache,
                warehouse_id=warehouse_id,
                customer_id=customer_id,
                purchase_date=purchase_date,
                invoice_no=invoice_no,
                handled_by=handled_by,
                import_source=import_source,
                import_items_summary=items_summary,
                import_total_qty=import_total_qty,
                note=_build_purchase_note(row, items_summary),
            )
            if purchase_result == "created":
                summary["purchases_created"] += 1
            elif purchase_result == "updated":
                summary["purchases_updated"] += 1
            else:
                summary["purchases_skipped_conflict"] += 1

        db.commit()
        summary["processed_warehouses"] = sorted(summary["processed_warehouses"])
        return summary
    finally:
        workbook.close()
